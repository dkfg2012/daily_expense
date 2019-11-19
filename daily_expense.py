import openpyxl
import datetime
import time
import calendar

column = ['transport', 'breakfast', 'lunch', 'dinner', 'snack', 'drink', 'entertainment', 'others (specific)']
day_of_week = {"Sun":"B", "Mon":"C", "Tue":"D", "Wed":"E", "Thu":"F", "Fri":"G", "Sat":"H"}
wb = openpyxl.load_workbook("expense_table.xlsx")
ws = wb.active
sheet = wb['Sheet1']


def today_weekday():
    today = datetime.datetime.now().strftime("%d-%m-%Y")
    weekday = tuple(list(map(int, datetime.datetime.now().strftime("%Y,%m,%d").split(','))))
    weekday = calendar.day_name[calendar.weekday(*weekday)][0:3]
    return today, weekday

def yesterday_weekday():
    yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
    yesterday = tuple(list(map(int, yesterday.strftime("%Y,%m,%d").split(','))))
    y_weekday = calendar.day_name[calendar.weekday(*yesterday)][0:3]
    return y_weekday

def current_row():
    if(sheet["Z1"].value != None):
        current_row = sheet["Z1"].value
        y_weekday = yesterday_weekday()
        try:
            if(y_weekday == "Sat"):
                Sat_row = list(ws.columns)[7]
                for cell in Sat_row:
                    if (cell.value != None or ws.max_row <= cell.row):
                        current_row = cell.row + 5
                    else:
                        current_row = ws.max_row + 5
            else:
                pass
        except:
            pass
    else:
        current_row = 1
    sheet["Z1"].value = current_row
    return current_row

today, weekday = today_weekday()

today_cell = day_of_week[weekday] + str(current_row())
transport_cell = day_of_week[weekday] + str(current_row() + 1)
breakfast_cell = day_of_week[weekday] + str(current_row() + 2)
lunch_cell = day_of_week[weekday] + str(current_row() + 3)
dinner_cell = day_of_week[weekday] + str(current_row() + 4)
snack_cell = day_of_week[weekday] + str(current_row() + 5)
drink_cell = day_of_week[weekday] + str(current_row() + 6)
entertainment_cell = day_of_week[weekday] + str(current_row() + 7)
other_cell = day_of_week[weekday] + str(current_row() + 8)

if (weekday == "Sat"):
    if(sheet[today_cell].value != None):
        sheet[today_cell].value = None
        today_cell = day_of_week[weekday] + str(current_row())
        transport_cell = day_of_week[weekday] + str(current_row() + 1)
        breakfast_cell = day_of_week[weekday] + str(current_row() + 2)
        lunch_cell = day_of_week[weekday] + str(current_row() + 3)
        dinner_cell = day_of_week[weekday] + str(current_row() + 4)
        snack_cell = day_of_week[weekday] + str(current_row() + 5)
        drink_cell = day_of_week[weekday] + str(current_row() + 6)
        entertainment_cell = day_of_week[weekday] + str(current_row() + 7)
        other_cell = day_of_week[weekday] + str(current_row() + 8)

def test_input(number):
    try:
        int(number)
    except Exception:
        return test_input(input("Enter must be integer, please enter again "))
    return number



def input_expense():
    sheet[today_cell] = today
    sheet[transport_cell] = "$ " + test_input(input("Your transport expense "))
    sheet[breakfast_cell] = "$ " + test_input(input("Your breakfast expense "))
    sheet[lunch_cell] = "$ " + test_input(input("Your lunch expense "))
    sheet[dinner_cell] = "$ " + test_input(input("Your dinner expense "))
    sheet[snack_cell] = "$ " + test_input(input("Your snack expense "))
    sheet[drink_cell] = "$ " + test_input(input("Your drink expense "))
    sheet[entertainment_cell] = "$ " + test_input(input("Your entertainmant expense "))
    sheet[other_cell] = "$ " + test_input(input("Your other expense "))

input_expense()

wb.save("expense_table.xlsx")